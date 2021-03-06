import sys
import datetime
from tactic_app import app, db, fs, socketio
from flask import request, jsonify, render_template, send_file, url_for
from flask_login import current_user, login_required
from flask_socketio import join_room
from tactic_app.main import delete_mainwindow
from tactic_app.shared_dicts import tile_classes, user_tiles, loaded_user_modules
from tactic_app.shared_dicts import mainwindow_instances, distribute_event, create_initial_metadata, get_tile_class
from user_manage_views import project_manager, collection_manager
import openpyxl
import cStringIO
import cPickle
from bson.binary import Binary


# The main window should join a room associated with the user
@socketio.on('connect', namespace='/main')
def connected_msg():
    print"client connected"


@socketio.on('join', namespace='/main')
def on_join(data):
    room = data["room"]
    join_room(room)
    print "user joined room " + room

# Views for creating and saving a new project
# As well as for updating an existing project.


@app.route('/save_new_project', methods=['POST'])
@login_required
def save_new_project():
    data_dict = request.json
    # noinspection PyBroadException
    try:
        mainwindow_instances[data_dict['main_id']].project_name = data_dict["project_name"]
        mainwindow_instances[data_dict['main_id']].hidden_columns_list = data_dict["hidden_columns_list"]
        mainwindow_instances[data_dict['main_id']].console_html = data_dict["console_html"]
        tspec_dict = data_dict["tablespec_dict"]
        for (dname, spec) in tspec_dict.items():
            mainwindow_instances[data_dict['main_id']].doc_dict[dname].table_spec = spec

        mainwindow_instances[data_dict['main_id']].loaded_modules = list(loaded_user_modules[current_user.username])
        project_dict = mainwindow_instances[data_dict['main_id']].compile_save_dict()
        save_dict = {}
        save_dict["metadata"] = create_initial_metadata()
        save_dict["project_name"] = project_dict["project_name"]
        save_dict["file_id"] = fs.put(Binary(cPickle.dumps(project_dict)))
        mainwindow_instances[data_dict['main_id']].mdata = save_dict["metadata"]

        db[current_user.project_collection_name].insert_one(save_dict)

        project_manager.update_selector_list()
        return jsonify({"project_name": data_dict["project_name"],
                        "success": True,
                        "message": "Project Successfully Saved"})
    except:
        mainwindow_instances[data_dict['main_id']].handle_exception("Error saving new project")
        return jsonify({"success": False})


@app.route("/add_blank_console_text/<main_id>", methods=["GET"])
@login_required
def add_blank_console_text(main_id):
    try:
        mainwindow_instances[main_id].print_to_console("<div contenteditable='true'></div>")
        return jsonify({"success": True})
    except:
        return jsonify({"success": False, "message": "Error creating console text area"})


@app.route("/send_log_html/<main_id>", methods=["POST"])
@login_required
def send_log_html(main_id):
    try:
        console_html = request.json["console_html"]
        mainwindow_instances[main_id].console_html = console_html
        return jsonify({"success": True})
    except:
        return jsonify({"success": False, "message": "Error opening log wndow"})

@app.route("/open_log_window/<main_id>", methods=["GET", "POST"])
@login_required
def open_log_window(main_id):
    if mainwindow_instances[main_id].project_name == None:
        title = mainwindow_instances[main_id].short_collection_name + " log"
    else:
        title = mainwindow_instances[main_id].project_name + " log"
    console_html = mainwindow_instances[main_id].console_html.decode("utf-8", "ignore").encode("ascii")
    return render_template("log_window_template.html", window_title=title, console_html=console_html)

@app.route('/update_project', methods=['POST'])
@login_required
def update_project():
    data_dict = request.json
    try:
        tspec_dict = data_dict["tablespec_dict"]
        mainwindow_instances[data_dict['main_id']].hidden_columns_list = data_dict["hidden_columns_list"]
        mainwindow_instances[data_dict['main_id']].console_html = data_dict["console_html"]
        mainwindow_instances[data_dict['main_id']].loaded_modules = list(loaded_user_modules[current_user.username])
        for (dname, spec) in tspec_dict.items():
            mainwindow_instances[data_dict['main_id']].doc_dict[dname].table_spec = spec
        project_dict = mainwindow_instances[data_dict['main_id']].compile_save_dict()
        pname = project_dict["project_name"]
        mainwindow_instances[data_dict['main_id']].mdata["updated"] = datetime.datetime.today()

        new_file_id = fs.put(Binary(cPickle.dumps(project_dict)))

        # Here we are trying to deal with both old-style and new-style saves
        # If it appears the project was saved old-style, then we'll delete and recreate it.
        save_dict = db[current_user.project_collection_name].find_one({"project_name": pname})
        if "file_id" in save_dict:
            fs.delete(save_dict["file_id"])
            save_dict["project_name"] = pname
            save_dict["metadata"] = mainwindow_instances[data_dict['main_id']].mdata
            save_dict["file_id"] = new_file_id
            db[current_user.project_collection_name].update_one({"project_name": pname},
                                                    {'$set': save_dict})
        else:
            db[current_user.project_collection_name].delete_one({"project_name": pname})
            save_dict = {}
            save_dict["project_name"] = pname
            save_dict["metadata"] = mainwindow_instances[data_dict['main_id']].mdata
            save_dict["file_id"] = new_file_id
            db[current_user.project_collection_name].insert_one(save_dict)

        return jsonify({"success": True, "message": "Project Successfully Saved"})
    except:
        mainwindow_instances[data_dict['main_id']].handle_exception("Error saving project")
        return jsonify({"success": False})


@app.route('/export_data', methods=['POST'])
@login_required
def export_data():
    data_dict = request.json
    doc_dict = mainwindow_instances[data_dict['main_id']].doc_dict
    full_collection_name = current_user.build_data_collection_name(data_dict['export_name'])
    for docinfo in doc_dict.values():
        db[full_collection_name].insert_one({"name": docinfo.name,
                                             "data_rows": docinfo.data_rows,
                                             "header_list": docinfo.header_list})
    collection_manager.update_selector_list()
    return jsonify({"success": True, "message": "Data Successfully Exported"})


@app.route('/download_table/<main_id>/<new_name>', methods=['GET', 'POST'])
@login_required
def download_table(main_id, new_name):
    mw = mainwindow_instances[main_id]
    doc_info = mw.doc_dict[mw.visible_doc_name]
    data_rows = doc_info.all_sorted_data_rows
    header_list = doc_info.header_list
    str_io = cStringIO.StringIO()
    for header in header_list[:-1]:
        str_io.write(str(header) + ',')
    str_io.write(str(header_list[-1]) + '\n')
    for row in data_rows:
        for header in header_list[:-1]:
            str_io.write(str(row[header]) + ",")
        str_io.write(str(row[header_list[-1]]) + '\n')
    str_io.seek(0)
    return send_file(str_io,
                     attachment_filename=new_name,
                     as_attachment=True)


@app.route('/download_collection/<main_id>/<new_name>', methods=['GET', 'POST'])
@login_required
def download_collection(main_id, new_name):
    mw = mainwindow_instances[main_id]
    wb = openpyxl.Workbook()
    first = True
    for (doc_name, doc_info) in mw.doc_dict.items():
        if first:
            ws = wb.active
            ws.title = doc_name
            first = False
        else:
            ws = wb.create_sheet(title=doc_name)
        data_rows = doc_info.all_sorted_data_rows
        header_list = doc_info.header_list
        for c, header in enumerate(header_list, start=1):
            _ = ws.cell(row=1, column=c, value=header)
        for r, row in enumerate(data_rows, start=2):
            for c, header in enumerate(header_list, start=1):
                _ = ws.cell(row=r, column=c, value=row[header])
    virtual_notebook = openpyxl.writer.excel.save_virtual_workbook(wb)
    str_io = cStringIO.StringIO()
    str_io.write(virtual_notebook)
    str_io.seek(0)
    return send_file(str_io,
                     attachment_filename=new_name,
                     as_attachment=True)

# Views for reading data from the database and
# passing back to the client.


@app.route('/grab_data/<main_id>/<doc_name>', methods=['get'])
@login_required
def grab_data(main_id, doc_name):
    return jsonify({"doc_name": doc_name,
                    "is_shrunk": mainwindow_instances[main_id].is_shrunk,
                    "left_fraction": mainwindow_instances[main_id].left_fraction,
                    "data_rows": mainwindow_instances[main_id].doc_dict[doc_name].displayed_data_rows,
                    "background_colors": mainwindow_instances[main_id].doc_dict[doc_name].displayed_background_colors,
                    "header_list": mainwindow_instances[main_id].doc_dict[doc_name].header_list,
                    "is_last_chunk": mainwindow_instances[main_id].doc_dict[doc_name].is_last_chunk,
                    "is_first_chunk": mainwindow_instances[main_id].doc_dict[doc_name].is_first_chunk,
                    "max_table_size": mainwindow_instances[main_id].doc_dict[doc_name].max_table_size})


@app.route('/grab_chunk_with_row', methods=['get', 'post'])
@login_required
def grab_chunk_with_row():
    data_dict = request.json
    doc_name = data_dict["doc_name"]
    main_id = data_dict["main_id"]
    row_id = data_dict["row_id"]
    mainwindow_instances[main_id].doc_dict[doc_name].move_to_row(row_id)
    return jsonify({"doc_name": doc_name,
                    "left_fraction": mainwindow_instances[main_id].left_fraction,
                    "is_shrunk": mainwindow_instances[main_id].is_shrunk,
                    "data_rows": mainwindow_instances[main_id].doc_dict[doc_name].displayed_data_rows,
                    "background_colors": mainwindow_instances[main_id].doc_dict[doc_name].displayed_background_colors,
                    "header_list": mainwindow_instances[main_id].doc_dict[doc_name].header_list,
                    "is_last_chunk": mainwindow_instances[main_id].doc_dict[doc_name].is_last_chunk,
                    "is_first_chunk": mainwindow_instances[main_id].doc_dict[doc_name].is_first_chunk,
                    "max_table_size": mainwindow_instances[main_id].doc_dict[doc_name].max_table_size,
                    "actual_row": mainwindow_instances[main_id].doc_dict[doc_name].get_actual_row(row_id)})


@app.route('/grab_next_chunk/<main_id>/<doc_name>', methods=['get'])
@login_required
def grab_next_chunk(main_id, doc_name):
    step_amount = mainwindow_instances[main_id].doc_dict[doc_name].advance_to_next_chunk()
    return jsonify({"doc_name": doc_name,
                    "data_rows": mainwindow_instances[main_id].doc_dict[doc_name].displayed_data_rows,
                    "background_colors": mainwindow_instances[main_id].doc_dict[doc_name].displayed_background_colors,
                    "header_list": mainwindow_instances[main_id].doc_dict[doc_name].header_list,
                    "is_last_chunk": mainwindow_instances[main_id].doc_dict[doc_name].is_last_chunk,
                    "is_first_chunk": mainwindow_instances[main_id].doc_dict[doc_name].is_first_chunk,
                    "step_size": step_amount})


@app.route('/grab_previous_chunk/<main_id>/<doc_name>', methods=['get'])
@login_required
def grab_previous_chunk(main_id, doc_name):
    step_amount = mainwindow_instances[main_id].doc_dict[doc_name].go_to_previous_chunk()
    if step_amount is None:  # This shouldn't happen
        return
    return jsonify({"doc_name": doc_name,
                    "data_rows": mainwindow_instances[main_id].doc_dict[doc_name].displayed_data_rows,
                    "background_colors": mainwindow_instances[main_id].doc_dict[doc_name].displayed_background_colors,
                    "header_list": mainwindow_instances[main_id].doc_dict[doc_name].header_list,
                    "is_last_chunk": mainwindow_instances[main_id].doc_dict[doc_name].is_last_chunk,
                    "is_first_chunk": mainwindow_instances[main_id].doc_dict[doc_name].is_first_chunk,
                    "step_size": step_amount})


@app.route('/grab_project_data/<main_id>/<doc_name>', methods=['get'])
@login_required
def grab_project_data(main_id, doc_name):
    mw = mainwindow_instances[main_id]
    return jsonify({"doc_name": doc_name,
                    "is_shrunk": mainwindow_instances[main_id].is_shrunk,
                    "tile_ids": mw.tile_sort_list,
                    "left_fraction": mw.left_fraction,
                    "data_rows": mw.doc_dict[doc_name].displayed_data_rows,
                    "background_colors": mainwindow_instances[main_id].doc_dict[doc_name].displayed_background_colors,
                    "hidden_columns_list": mw.hidden_columns_list,
                    "is_last_chunk": mainwindow_instances[main_id].doc_dict[doc_name].is_last_chunk,
                    "is_first_chunk": mainwindow_instances[main_id].doc_dict[doc_name].is_first_chunk,
                    "tablespec_dict": mw.tablespec_dict()})


@app.route('/get_menu_template', methods=['get'])
@login_required
def get_menu_template():
    return send_file("templates/menu_template.html")


@app.route('/get_table_templates', methods=['get'])
@login_required
def get_table_templates():
    return send_file("templates/table_templates.html")


@app.route('/remove_mainwindow/<main_id>', methods=['post'])
@login_required
def remove_mainwindow(main_id):
    delete_mainwindow(main_id)
    return

@app.route('/get_tile_types', methods=['GET'])
@login_required
def get_tile_types():
    tile_types = {}
    for (category, the_dict) in tile_classes.items():
        tile_types[category] = the_dict.keys()

    if current_user.username in user_tiles:
        for (category, the_dict) in user_tiles[current_user.username].items():
            if category not in tile_types:
                tile_types[category] = []
            tile_types[category] += the_dict.keys()
    result = {"tile_types": tile_types}
    return jsonify(result)


@app.route('/set_visible_doc/<main_id>/<doc_name>', methods=['get', 'post'])
@login_required
def set_visible_doc(main_id, doc_name):
    mainwindow_instances[main_id].visible_doc_name = doc_name
    return jsonify({"success": True})


@app.route('/distribute_events/<event_name>', methods=['get', 'post'])
@login_required
def distribute_events_stub(event_name):
    data_dict = request.json
    main_id = request.json["main_id"]

    # If necessary, have to convert the row_index on the client side to the row_id
    if (data_dict is not None) and ("active_row_index" in data_dict) and ("doc_name" in data_dict):
        if data_dict["active_row_index"] is not None:
            mwindow = mainwindow_instances[main_id]
            data_dict["active_row_index"] = mwindow.doc_dict[data_dict["doc_name"]].get_id_from_actual_row(data_dict["active_row_index"])
    if "tile_id" in request.json:
        tile_id = request.json["tile_id"]
    else:
        tile_id = None
    success = distribute_event(event_name, main_id, data_dict, tile_id)
    return jsonify({"success": success})


@app.route('/figure_source/<main_id>/<tile_id>/<figure_name>', methods=['GET', 'POST'])
@login_required
def figure_source(main_id, tile_id, figure_name):
    img = mainwindow_instances[main_id].tile_instances[tile_id].img_dict[figure_name]
    img_file = cStringIO.StringIO()
    img_file.write(img)
    img_file.seek(0)
    return send_file(img_file, mimetype='image/png')


@app.route('/data_source/<main_id>/<tile_id>/<data_name>', methods=['GET'])
@login_required
def data_source(main_id, tile_id, data_name):
    try:
        the_data = mainwindow_instances[main_id].tile_instances[tile_id].data_dict[data_name]
        return jsonify({"success": True, "data": the_data})
    except:
        error_string = str(sys.exc_info()[0]) + " " + str(sys.exc_info()[1])
        mainwindow_instances[main_id].handle_exception("Error getting data " + error_string)
        return jsonify({"success": False})


# noinspection PyUnresolvedReferences
@app.route('/create_tile_request/<tile_type>', methods=['GET', 'POST'])
@login_required
def create_tile_request(tile_type):
    main_id = request.json["main_id"]
    # noinspection PyBroadException
    try:
        tile_name = request.json["tile_name"]
        new_tile = mainwindow_instances[main_id].create_tile_instance_in_mainwindow(tile_type, tile_name)
        tile_id = new_tile.tile_id
        form_html = new_tile.create_form_html()
        result = render_template("tile.html", tile_id=tile_id,
                                 tile_name=new_tile.tile_name,
                                 form_text=form_html)
        return jsonify({"success": True, "html": result, "tile_id": tile_id})
    except:
        error_string = str(sys.exc_info()[0]) + " " + str(sys.exc_info()[1])
        mainwindow_instances[main_id].handle_exception("Error creating tile " + error_string)
        return jsonify({"success": False})


@app.route('/reload_tile/<tile_id>', methods=['GET', 'POST'])
@login_required
def reload_tile(tile_id):
    save_attrs_for_reload = ["tile_id", "tile_name", "header_height", "front_height", "front_width", "back_height", "back_width",
                  "tda_width", "tda_height", "width", "height",
                  "full_tile_width", "full_tile_height", "is_shrunk", "configured"]
    try:
        main_id = request.json["main_id"]
        mw = mainwindow_instances[main_id]
        old_instance = mw.tile_instances[tile_id]
        old_instance.kill()
        saved_options = {}
        for option in old_instance.options:
            attr = option["name"]
            if hasattr(old_instance, attr):
                saved_options[attr] = getattr(old_instance, attr)
        tile_type = old_instance.tile_type
        tile_name = old_instance.tile_name
        new_cls = get_tile_class(current_user.username, tile_type)
        new_instance = new_cls(main_id, tile_id, tile_name)
        for attr, val in saved_options.items():
            setattr(new_instance, attr, val)
        for attr in save_attrs_for_reload:
            setattr(new_instance, attr, getattr(old_instance, attr))
        form_html = new_instance.create_form_html()
        mw.tile_instances[tile_id] = new_instance
        new_instance.start()
        return jsonify({"success": True, "html": form_html})
    except:
        error_string = str(sys.exc_info()[0]) + " " + str(sys.exc_info()[1])
        mainwindow_instances[main_id].handle_exception("Error reloading tile " + error_string)
        return jsonify({"success": False})

@app.route('/create_tile_from_save_request/<tile_id>', methods=['GET', 'POST'])
@login_required
def create_tile_from_save_request(tile_id):
    main_id = request.json["main_id"]
    # noinspection PyBroadException
    try:
        tile_instance = mainwindow_instances[main_id].tile_instances[tile_id]
        tile_instance.figure_url = url_for("figure_source", main_id=main_id, tile_id=tile_id, figure_name="X")[:-1]
        form_html = tile_instance.create_form_html()
        if tile_instance.is_shrunk:
            dsr_string = ""
            dbr_string = "display: none"
            bda_string = "display: none"
            main_height = tile_instance.header_height
        else:
            dsr_string = "display: none"
            dbr_string = ""
            bda_string = ""
            main_height = tile_instance.full_tile_height
        result = render_template("saved_tile.html", tile_id=tile_id,
                                 tile_name=tile_instance.tile_name,
                                 form_text=form_html,
                                 current_html=tile_instance.current_html,
                                 whole_width=tile_instance.full_tile_width,
                                 whole_height=main_height,
                                 front_height=tile_instance.front_height,
                                 front_width=tile_instance.front_width,
                                 back_height=tile_instance.back_height,
                                 back_width=tile_instance.back_width,
                                 tda_height=tile_instance.tda_height,
                                 tda_width=tile_instance.tda_width,
                                 is_strunk=tile_instance.is_shrunk,
                                 triangle_right_display_string=dsr_string,
                                 triangle_bottom_display_string=dbr_string,
                                 front_back_display_string=bda_string

                                 )
        return jsonify({"html": result, "tile_id": tile_id, "is_shrunk": tile_instance.is_shrunk,
                        "saved_size": tile_instance.full_tile_height})
    except:
        error_string = str(sys.exc_info()[0]) + " " + str(sys.exc_info()[1])
        mainwindow_instances[main_id].handle_exception("Error creating tile from save " + error_string)
        return jsonify({"success": False})
